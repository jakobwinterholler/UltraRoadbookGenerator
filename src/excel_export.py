"""Export climb data and gradient analysis to an Excel roadbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from climb_detector import Climb
from gradient_analysis import ClimbGradientStats

# Gradient columns receive colour-coded conditional formatting.
GRADIENT_COLUMNS = ("G", "H", "I", "J", "K", "L")

# Conditional-formatting colour bands (gradient %).
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_FILL_RED = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
_FILL_PURPLE = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")

HEADERS = (
    "Nickname",
    "ID",
    "Start km",
    "End km",
    "Length (km)",
    "Elevation Gain (m)",
    "Average Gradient (%)",
    "Max 50 m (%)",
    "Max 100 m (%)",
    "Max 250 m (%)",
    "Max 500 m (%)",
    "Max 1000 m (%)",
)


def _apply_gradient_formatting(ws, row_count: int) -> None:
    """Apply colour bands to all gradient columns for data rows."""
    if row_count < 2:
        return

    for column in GRADIENT_COLUMNS:
        cell_range = f"{column}2:{column}{row_count}"
        anchor = f"{column}2"

        # Rules are evaluated top-to-bottom; first match wins.
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{anchor}<4"], fill=_FILL_GREEN),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({anchor}>=4,{anchor}<6)"], fill=_FILL_YELLOW),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({anchor}>=6,{anchor}<8)"], fill=_FILL_ORANGE),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"AND({anchor}>=8,{anchor}<10)"], fill=_FILL_RED),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f"{anchor}>=10"], fill=_FILL_PURPLE),
        )


def _autofit_columns(ws) -> None:
    """Set each column width to fit its content."""
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = max_length + 2


def export_roadbook(
    climbs_with_gradients: list[tuple[Climb, ClimbGradientStats]],
    output_path: Path,
    *,
    nicknames: dict[str, str] | None = None,
) -> None:
    """
    Write climb data and gradient statistics to an Excel roadbook.

    Applies bold/frozen header, auto-fitted columns, and gradient colour bands.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Roadbook"

    ws.append(list(HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    resolved_nicknames = nicknames or {}

    for climb, gradients in climbs_with_gradients:
        ws.append(
            [
                resolved_nicknames.get(climb.climb_id) or f"Climb {climb.climb_id.replace('C', '')}",
                climb.climb_id,
                round(climb.start_km, 2),
                round(climb.end_km, 2),
                round(climb.length_km, 2),
                round(climb.elevation_gain_m),
                round(climb.avg_gradient_pct, 1),
                gradients.max_50_m_pct,
                gradients.max_100_m_pct,
                gradients.max_250_m_pct,
                gradients.max_500_m_pct,
                gradients.max_1000_m_pct,
            ]
        )

    ws.freeze_panes = "A2"
    _apply_gradient_formatting(ws, ws.max_row)
    _autofit_columns(ws)

    wb.save(output_path)
